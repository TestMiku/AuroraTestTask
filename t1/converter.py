import logging
import tkinter as tk
import tkinter.filedialog
from tkinter import messagebox
from docx2pdf import convert
from openpyxl import load_workbook
from PyPDF2 import PdfWriter
from reportlab.lib.pagesizes import letter, landscape
from reportlab.pdfgen import canvas

# Настройка логирования
logging.basicConfig(filename='conversion.log', level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')

def convert_to_pdf(input_file):
    try:
        # Проверяем формат входного файла и осуществляем соответствующую конвертацию
        if input_file.endswith('.docx'):
            output_file = input_file.replace('.docx', '.pdf')
            convert(input_file, output_file)
            # Записываем успешное завершение операции в лог
            logging.info(f'Конвертация {input_file} в {output_file} прошла успешно.')
            # Выводим сообщение об успешном завершении операции для пользователя
            messagebox.showinfo("Успех", f"Документ успешно сконвертирован в PDF: {output_file}")
        elif input_file.endswith('.xlsx'):
            wb = load_workbook(input_file)
            sheet = wb.active
            max_row = sheet.max_row
            max_column = sheet.max_column
            output_file = input_file.replace('.xlsx', '.pdf')
            c = canvas.Canvas(output_file, pagesize=landscape(letter))

            x_offset = 50  # Смещение по горизонтали
            y_offset = 600  # Смещение по вертикали

            # Проходим по всем ячейкам листа Excel и рисуем их содержимое на PDF-документе
            for row in range(1, max_row + 1):
                for column in range(1, max_column + 1):
                    cell = sheet.cell(row=row, column=column)
                    c.drawString(x_offset + (column - 1) * 100, y_offset - (row - 1) * 12, str(cell.value))
            
            c.save()

            # Записываем успешное завершение операции в лог
            logging.info(f'Конвертация {input_file} в {output_file} прошла успешно.')
            # Выводим сообщение об успешном завершении операции для пользователя
            messagebox.showinfo("Успех", f"Документ успешно сконвертирован в PDF: {output_file}")
        else:
            # Если формат файла не .docx и не .xlsx, вызываем исключение
            raise ValueError("Неподдерживаемый формат файла.")
    except Exception as e:
        # Обрабатываем любые ошибки, которые могут возникнуть в процессе конвертации
        logging.error(f'Ошибка конвертации {input_file}: {str(e)}')
        # Выводим сообщение об ошибке для пользователя
        messagebox.showerror("Ошибка", f"Ошибка конвертации: {str(e)}")

def main():
    # Создаем графическое окно для выбора файла
    root = tk.Tk()
    root.withdraw()
    input_file = tk.filedialog.askopenfilename(title="Выберите файл для конвертации")
    if input_file:
        # Если выбран файл, запускаем процесс конвертации
        convert_to_pdf(input_file)

if __name__ == "__main__":
    main()
